from rest_framework import viewsets, filters, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
import io
import csv
from datetime import datetime
import openpyxl

from .models import Client, Team, Segment, Project
from .serializers import ClientSerializer, TeamSerializer, SegmentSerializer, ProjectSerializer

# Colunas obrigatórias (mínimo para criar um projeto)
REQUIRED_COLUMNS = {'Cliente', 'Projeto'}

# Colunas esperadas completas (para informação no modal, não bloqueantes)
FULL_COLUMNS = {'Cliente', 'Projeto', 'Término Real', 'Total', 'Segmento', 'Time', 'Squad', 'Impetitivos', 'Status'}

def validate_headers(headers):
    """Verifica se as colunas mínimas obrigatórias estão presentes."""
    present = {h.strip() for h in headers if h}
    missing = REQUIRED_COLUMNS - present
    return sorted(missing)

def parse_row(row_dict):
    """Converte um dicionário de linha (CSV ou XLSX) para campos do projeto."""
    impediments_val = (row_dict.get('Impetitivos') or row_dict.get('Impeditivos') or '').strip()
    return {
        'client_name': (row_dict.get('Cliente') or '').strip(),
        'proj_name': (row_dict.get('Projeto') or '').strip(),
        'date_str': str(row_dict.get('Término Real') or '').strip(),
        'total_val': str(row_dict.get('Total') or '0').strip(),
        'segment_name': (row_dict.get('Segmento') or '').strip(),
        'team_name': (row_dict.get('Time') or '').strip(),
        'impeditivos': impediments_val,
        'status_raw': (row_dict.get('Status') or '').strip().lower(),
        'squad': (row_dict.get('Squad') or '').strip(),
    }

def save_project(data):
    """Cria ou atualiza um projeto com base nos dados da linha. Retorna True se bem-sucedido."""
    if not data['client_name'] or not data['proj_name']:
        return False

    from .models import Project as ProjectModel
    status_code = ProjectModel.STATUS_MAP.get(data['status_raw'], 'EM_ANDAMENTO')

    client, _ = Client.objects.get_or_create(name=data['client_name'])
    team = Team.objects.get_or_create(name=data['team_name'])[0] if data['team_name'] else None
    segment = Segment.objects.get_or_create(name=data['segment_name'])[0] if data['segment_name'] else None

    actual_date = None
    ds = data['date_str']
    if ds and ds not in ('-', 'None', ''):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
            try:
                actual_date = datetime.strptime(ds, fmt).date()
                break
            except ValueError:
                continue

    # Total = quantidade de entregas (pode ser 0)
    try:
        delivery_count = int(float(data['total_val'])) if data['total_val'] else 0
    except (ValueError, TypeError):
        delivery_count = 0

    Project.objects.update_or_create(
        name=data['proj_name'],
        client=client,
        defaults={
            'actual_end_date': actual_date,
            'delivery_count': delivery_count,
            'segment': segment,
            'team': team,
            'impediments': data['impeditivos'],
            'squad': data.get('squad', ''),
            'status': status_code,
        }
    )
    return True

class DashboardViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API for the main dashboard metrics.
    """
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer

    @action(detail=False, methods=['get'])
    def metrics(self, request):
        total_projects = Project.objects.count()
        by_team = {}
        for team in Team.objects.all():
            by_team[team.name] = Project.objects.filter(team=team).count()
        
        by_segment = {}
        for seg in Segment.objects.all():
            by_segment[seg.name] = Project.objects.filter(segment=seg).count()

        return Response({
            'total_projects': total_projects,
            'by_team': by_team,
            'by_segment': by_segment
        })

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all().order_by('-actual_end_date')
    serializer_class = ProjectSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'client__name', 'team__name', 'segment__name']

    @action(detail=False, methods=['delete'])
    def clear_all(self, request):
        total, _ = Project.objects.all().delete()
        return Response({
            'message': f'{total} projeto(s) excluído(s) com sucesso.'
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def import_file(self, request):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'Nenhum arquivo enviado.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = uploaded_file.name.lower()
        count = 0
        errors = []
        rows = []

        try:
            if filename.endswith('.csv'):
                decoded = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                for row in reader:
                    rows.append(dict(row))

            elif filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

                # Campos que usam fill-down (quando vazio, herda da linha anterior)
                FILL_DOWN_FIELDS = ['Cliente', 'Segmento', 'Time', 'Squad', 'Status']
                last_values = {f: '' for f in FILL_DOWN_FIELDS}

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)}

                    # Pula linhas completamente vazias
                    if not any(row_dict.values()):
                        continue

                    # Aplica fill-down: herda valor anterior quando célula está vazia
                    for field in FILL_DOWN_FIELDS:
                        if field in row_dict:
                            if row_dict[field]:
                                last_values[field] = row_dict[field]
                            else:
                                row_dict[field] = last_values[field]

                    rows.append(row_dict)
            else:
                return Response({'error': 'Formato inválido. Envie um arquivo .csv ou .xlsx.'}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({'error': f'Erro ao ler o arquivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Validar colunas obrigatórias
        if rows:
            missing = validate_headers(rows[0].keys())
            if missing:
                return Response({
                    'error': f'Colunas obrigatórias ausentes na planilha: {", ".join(missing)}. '
                             f'Certifique-se de que o arquivo contém: Cliente, Projeto, Término Real, Total, Segmento, Time, Impetitivos, Status.'
                }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for row in rows:
                try:
                    data = parse_row(row)
                    if save_project(data):
                        count += 1
                except Exception as e:
                    errors.append(f'Linha "{row.get("Projeto", "?")}": {str(e)}')

        return Response({
            'message': f'{count} projeto(s) importado(s) com sucesso.',
            'errors': errors
        }, status=status.HTTP_201_CREATED)
